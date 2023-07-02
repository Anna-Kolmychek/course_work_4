import json
import os.path

from openpyxl import Workbook, styles, load_workbook

from parser_vacancies.classes.file_handler import FileHandler
from parser_vacancies.classes.vacancies_handler import VacanciesHandler
from parser_vacancies.classes.vacancy import Vacancy
from parser_vacancies.constants import DATA_DIR


class ExcelFileHandler(FileHandler):
    def __init__(self, file_name: str):
        """инициализация экземпляра через имя файла.
        добавляется путь к папке с файлами и расширение файла"""
        super().__init__(file_name)
        self.file_name = os.path.join(DATA_DIR, f'{file_name}.xlsx')
        if self.is_file_exist():
            self.wb = load_workbook(self.file_name)
        else:
            self.wb = Workbook()
        self.ws = self.wb.active

    def overwrite_vacancies(self, vacancies: VacanciesHandler) -> None:
        """перезаписывает данные в файл (старое содержимое удаляет, новое записывает)"""
        sheet_name = self.wb.sheetnames[0]
        del self.wb[sheet_name]
        self.ws = self.wb.create_sheet(sheet_name)

        self.prepare_new_file()
        self.add_vacancies(vacancies)

    def add_vacancies(self, vacancies: VacanciesHandler) -> None:
        """дозаписывает данные в файл (старое содержимое НЕ удаляет, новое записывает в конец)"""
        prepared_data = self.convert_vacancies_to_data(vacancies)
        align_center = styles.Alignment(horizontal='center')
        for item in prepared_data:
            self.ws.append(item)
            self.ws.cell(row=self.ws._current_row, column=5).alignment = align_center
            self.ws.cell(row=self.ws._current_row, column=6).alignment = align_center
            self.ws.cell(row=self.ws._current_row, column=7).alignment = align_center
            self.ws.cell(row=self.ws._current_row, column=8).alignment = align_center
        try:
            self.wb.save(self.file_name)
        except PermissionError:
            print('-' * 10,
                  'Доступ к файлу ограничен. Проверьте, закрыт ли файл или задайте другое имя файла.',
                  '-' * 10)

    def read_vacancies(self) -> VacanciesHandler:
        """считывает вакансии из файла"""
        data_from_file = []
        for item in self.ws.values:
            data_from_file.append(list(item))
        data_from_file.pop(0)
        return self.convert_data_to_vacancies(data_from_file)

    @staticmethod
    def convert_data_to_vacancies(data) -> VacanciesHandler:
        """переводит данные из формата, полученного из файла, в формат VacanciesHandler"""
        vacancies = VacanciesHandler([])
        for item in data:
            if None in [item[0], item[1], item[2], item[3], item[6], item[7], item[8]]:
                print(f'Данные вакансии {item[0]} не полные и не могут быть считаны из файла')
            else:
                if item[6] == '+':
                    distant_work = True
                else:
                    distant_work = False
                vacancy = Vacancy(item[0],
                                  item[1],
                                  item[2],
                                  item[3],
                                  item[4],
                                  item[5],
                                  distant_work,
                                  item[7],
                                  item[8])
                vacancies.append(vacancy)
        return vacancies

    @staticmethod
    def convert_vacancies_to_data(vacancies: VacanciesHandler) -> list:
        """переводит данные из формата VacanciesHandler в формат для записи в файл"""
        prepared_data = []
        for vacancy in vacancies:
            if vacancy.distant_work:
                distant_work = '+'
            else:
                distant_work = '-'
            prepared_data.append([vacancy.vacancy_id,
                                  vacancy.title,
                                  vacancy.url,
                                  vacancy.description,
                                  vacancy.payment_from,
                                  vacancy.payment_to,
                                  distant_work,
                                  vacancy.date_published,
                                  vacancy.town])
        return prepared_data

    def prepare_new_file(self):
        # Устанавливаем ширину колонок
        self.ws.column_dimensions['A'].width = 12
        self.ws.column_dimensions['B'].width = 45
        self.ws.column_dimensions['C'].width = 15
        self.ws.column_dimensions['D'].width = 45
        self.ws.column_dimensions['E'].width = 12
        self.ws.column_dimensions['F'].width = 12
        self.ws.column_dimensions['G'].width = 10
        self.ws.column_dimensions['H'].width = 17
        self.ws.column_dimensions['I'].width = 17

        # Задаем стили
        font_bold = styles.Font(bold=True)
        align_center = styles.Alignment(horizontal='center')

        # Задаем заголовки со стилями
        self.ws.append(['ID',
                        'Название',
                        'Ссылка',
                        'Описание (фрагмент)',
                        'ЗП от',
                        'ЗП до',
                        'Удаленка',
                        'Дата публикации',
                        'Место работы'])
        for cell in self.ws[1:1]:
            cell.font = font_bold
            cell.alignment = align_center

    def is_file_exist(self) -> bool:
        """проверяет, есть ли файл с таким именем"""
        return os.path.isfile(self.file_name)

    def delete_file(self) -> None:
        """удаляет файл"""
        if self.is_file_exist():
            os.remove(self.file_name)


# РАБОЧЕЕ ДЛЯ ПРОВЕРКИ
vacancy1 = Vacancy('hhvacancy_id1', 'title1', 'url1', 'description1',
                   20, 40, False,
                   '2023-06-25', 'town1')

vacancy2 = Vacancy('sjvacancy_id2', 'title2', 'url2', 'description2',
                   None, 50, False,
                   '2023-06-27', 'town2')

vacancy3 = Vacancy('sjvacancy_id3', 'tile3', 'url3', 'description3',
                   None, None, True,
                   '2023-06-26', 'town3')

vacancies = VacanciesHandler([vacancy1, vacancy2, vacancy3])
excel_fh = ExcelFileHandler('test3')
# excel_fh.read_vacancies()
# json_fh.overwrite_vacancies(vacancies)
# vacancies2 = VacanciesHandler(json_fh.read_vacancies())
# print(vacancies2)
